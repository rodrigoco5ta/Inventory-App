
import copy
import io
import os
from collections import defaultdict
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

st.set_page_config(page_title="Inventário de Localizações", layout="wide")

GREEN_FILL = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
YELLOW_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")

REQUIRED_HEADERS = [
    "Storage Bin",
    "Quantity",
    "Handling Unit",
    "Product",
    "Product Description",
]


def init_state():
    defaults = {
        "source_name": "",
        "source_bytes": None,
        "original_data": {},
        "closed_locations": set(),
        "reports": [],
        "current_location": None,
        "current_articles": {},
        "current_product": None,
        "hu_article_open": None,
        "location_input": "",
        "file_loaded": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def to_text(value):
    return "" if value is None else str(value).strip()


def to_number(value):
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            return 0.0


def fmt_qty(value):
    if value is None:
        return ""
    value = float(value or 0)
    if abs(value - int(value)) < 1e-9:
        return str(int(value))
    return f"{value:.3f}".rstrip("0").rstrip(".")


def hus_text(hus_dict):
    return " | ".join(
        f"{hu}={fmt_qty(qty)}" for hu, qty in sorted(hus_dict.items(), key=lambda x: x[0])
    )


def col_letter(idx):
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def read_original_data_from_bytes(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if "Data" not in wb.sheetnames:
        raise ValueError("A folha 'Data' não existe no ficheiro.")
    ws = wb["Data"]
    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError("Faltam colunas obrigatórias na folha Data: " + ", ".join(missing))

    by_loc = defaultdict(dict)
    col_bin = headers["Storage Bin"]
    col_qty = headers["Quantity"]
    col_hu = headers["Handling Unit"]
    col_prod = headers["Product"]
    col_desc = headers["Product Description"]

    for r in range(2, ws.max_row + 1):
        loc = to_text(ws.cell(r, col_bin).value)
        if not loc:
            continue
        prod = to_text(ws.cell(r, col_prod).value)
        desc = to_text(ws.cell(r, col_desc).value)
        hu = to_text(ws.cell(r, col_hu).value)
        qty = to_number(ws.cell(r, col_qty).value)

        if prod not in by_loc[loc]:
            by_loc[loc][prod] = {
                "description": desc,
                "original_total": 0.0,
                "counted_total": 0.0,
                "done": False,
                "hus_closed": False,
                "hus_initialized": False,
                "original_hus": defaultdict(float),
                "counted_hus": {},
            }
        rec = by_loc[loc][prod]
        rec["original_total"] += qty
        if hu:
            rec["original_hus"][hu] += qty

    cleaned = {}
    for loc, products in by_loc.items():
        cleaned[loc] = {}
        for prod, rec in products.items():
            rec["original_hus"] = dict(rec["original_hus"])
            rec["counted_hus"] = {}
            rec["counted_total"] = 0.0
            cleaned[loc][prod] = rec
    return cleaned


def reset_current_view(clear_location=True):
    st.session_state.current_location = None
    st.session_state.current_articles = {}
    st.session_state.current_product = None
    st.session_state.hu_article_open = None
    if clear_location:
        st.session_state.location_input = ""


def load_location(loc: str):
    original_data = st.session_state.original_data
    if not original_data:
        raise ValueError("Carrega primeiro o ficheiro de stocks.")
    if st.session_state.hu_article_open:
        rec = st.session_state.current_articles.get(st.session_state.hu_article_open, {})
        if rec and rec.get("hus_closed") is False:
            raise ValueError("Fecha primeiro as HUs do artigo em edição antes de trocar de localização.")
    loc = loc.strip()
    if not loc:
        raise ValueError("Introduz a localização.")
    if loc not in original_data:
        raise ValueError(f"A localização '{loc}' não existe na folha Data.")

    reset_current_view(clear_location=False)
    st.session_state.current_location = loc
    st.session_state.current_articles = copy.deepcopy(original_data[loc])
    products = sorted(st.session_state.current_articles.keys())
    st.session_state.current_product = products[0] if products else None


def ensure_article_open(product: str):
    rec = st.session_state.current_articles[product]
    if st.session_state.hu_article_open and st.session_state.hu_article_open != product:
        current_open = st.session_state.current_articles[st.session_state.hu_article_open]
        if not current_open["hus_closed"]:
            raise ValueError(f"Ainda tens as HUs do artigo {st.session_state.hu_article_open} em aberto.")
    if not rec["hus_initialized"]:
        rec["counted_hus"] = {hu: None for hu in rec["original_hus"].keys()}
        rec["counted_total"] = 0.0
        rec["hus_initialized"] = True
        rec["hus_closed"] = False
        rec["done"] = False
    st.session_state.current_product = product
    st.session_state.hu_article_open = product


def recalc_article(product: str):
    rec = st.session_state.current_articles[product]
    rec["counted_total"] = sum(float(v or 0) for v in rec["counted_hus"].values() if v is not None)


def close_hus_for_current_article():
    product = st.session_state.hu_article_open
    if not product:
        raise ValueError("Abre primeiro as HUs de um artigo.")
    rec = st.session_state.current_articles[product]
    if rec["counted_hus"]:
        missing = [hu for hu, qty in rec["counted_hus"].items() if qty is None]
        if missing:
            raise ValueError("Ainda existem HUs com quantidade contada em branco.")
        recalc_article(product)
    else:
        rec["counted_total"] = 0.0
    rec["hus_closed"] = True
    rec["done"] = False


def complete_article(product: str):
    rec = st.session_state.current_articles[product]
    if not rec["hus_closed"]:
        raise ValueError("Só podes concluir o artigo depois de fechar as HUs desse artigo.")
    rec["done"] = True
    if st.session_state.hu_article_open == product:
        st.session_state.hu_article_open = None


def next_open_article(current_product: str):
    products = [p for p in sorted(st.session_state.current_articles.keys()) if not st.session_state.current_articles[p]["done"]]
    if not products:
        return None
    if current_product not in products:
        return products[0]
    idx = products.index(current_product)
    return products[idx + 1] if idx + 1 < len(products) else None


def added_hus_text(rec):
    added = []
    for hu, qty in sorted(rec["counted_hus"].items()):
        if qty is None or abs(float(qty)) <= 1e-9:
            continue
        if hu not in rec["original_hus"]:
            added.append(f"{hu} ({fmt_qty(qty)})")
    return ", ".join(added) if added else ""


def action_with_hu_creation(base_action, rec):
    added_text = added_hus_text(rec)
    return f"{base_action} | Criar HUs: {added_text}" if added_text else base_action


def classify_article(rec):
    orig_total = round(float(rec["original_total"]), 6)
    counted_total = round(float(rec["counted_total"]), 6)
    orig_hus = {k: round(float(v), 6) for k, v in rec["original_hus"].items() if abs(v) > 1e-9}
    counted_hus = {k: round(float(v), 6) for k, v in rec["counted_hus"].items() if v is not None and abs(v) > 1e-9}
    same_codes = set(orig_hus.keys()) == set(counted_hus.keys())
    same_structure = same_codes and all(orig_hus.get(k, 0.0) == counted_hus.get(k, 0.0) for k in set(orig_hus) | set(counted_hus))
    added_hus = added_hus_text(rec)

    if same_structure and orig_total == counted_total:
        notes = "Sem ação corretiva."
        if added_hus:
            notes += f" Criar HUs: {added_hus}."
        return "Quantidade e HUs corretas", action_with_hu_creation("Prosseguir para o artigo/localização seguinte", rec), notes
    if same_codes and counted_total > orig_total:
        notes = f"Excesso identificado: {fmt_qty(counted_total - orig_total)}."
        if added_hus:
            notes += f" Criar HUs: {added_hus}."
        return "HUs corretas e quantidade a mais", action_with_hu_creation("Movimentar a quantidade a mais para uma localização fora do armazém (ex.: SPARE)", rec), notes
    if same_codes and counted_total < orig_total:
        notes = f"Falta identificada: {fmt_qty(orig_total - counted_total)}."
        if added_hus:
            notes += f" Criar HUs: {added_hus}."
        return "HUs corretas e quantidade a menos", action_with_hu_creation("Verificar localizações spare e movimentar quantidade e/ou puxar quantidade de localizações mais à frente do corredor", rec), notes
    if counted_total == orig_total and not same_structure:
        notes = "O total bate certo, mas a estrutura das HUs não coincide com a extração."
        if added_hus:
            notes += f" Criar HUs: {added_hus}."
        return "Quantidade total correta mas material mal identificado", action_with_hu_creation("Fazer repack e etiquetar devidamente o material", rec), notes
    notes = "Existem diferenças simultâneas de total e estrutura de HUs."
    if added_hus:
        notes += f" Criar HUs: {added_hus}."
    return "Diferença mista / revisão manual", action_with_hu_creation("Rever fisicamente o artigo, confirmar HUs e decidir ajuste em SAP manualmente", rec), notes


def close_location():
    current_location = st.session_state.current_location
    if not current_location:
        raise ValueError("Carrega primeiro uma localização.")
    if st.session_state.hu_article_open and not st.session_state.current_articles[st.session_state.hu_article_open]["done"]:
        raise ValueError("Ainda tens um artigo em edição. Fecha-o antes de fechar a localização.")
    not_done = [p for p, rec in st.session_state.current_articles.items() if not rec["done"]]
    if not not_done and not st.session_state.current_articles:
        raise ValueError("Não há artigos carregados para fechar.")
    if not_done:
        raise ValueError("Só podes fechar a localização depois de fechar cada artigo.")

    st.session_state.reports = [r for r in st.session_state.reports if r["Storage Bin"] != current_location]
    for product, rec in sorted(st.session_state.current_articles.items(), key=lambda x: x[0]):
        result, action, notes = classify_article(rec)
        row = {
            "Storage Bin": current_location,
            "Product": product,
            "Qtd Original": rec["original_total"],
            "Qtd Contada": rec["counted_total"],
            "HUs Originais": hus_text(rec["original_hus"]),
            "HUs Contadas": hus_text({k: v for k, v in rec["counted_hus"].items() if v is not None}),
            "HUs a Criar": added_hus_text(rec),
            "Resultado": result,
            "Acao SAP": action,
            "Observacoes": notes,
            "Data Fecho": datetime.now(),
        }
        st.session_state.reports.append(row)

    st.session_state.closed_locations.add(current_location)


def ensure_output_columns(ws, headers):
    wanted = ["Inventariado", "Data Fecho", "Resumo SAP", "Observacoes Inventario"]
    next_col = ws.max_column + 1
    for name in wanted:
        if name not in headers:
            ws.cell(1, next_col).value = name
            ws.cell(1, next_col).fill = YELLOW_FILL
            headers[name] = next_col
            next_col += 1


def write_reports_sheet(wb, reports):
    if "Relatorio" in wb.sheetnames:
        del wb["Relatorio"]
    wsr = wb.create_sheet("Relatorio")
    cols = [
        "Storage Bin", "Product", "Qtd Original", "Qtd Contada", "HUs Originais", "HUs Contadas",
        "HUs a Criar", "Resultado", "Acao SAP", "Observacoes", "Data Fecho",
    ]
    for c, name in enumerate(cols, start=1):
        cell = wsr.cell(1, c)
        cell.value = name
        cell.fill = YELLOW_FILL
    for r_idx, row in enumerate(reports, start=2):
        for c_idx, name in enumerate(cols, start=1):
            cell = wsr.cell(r_idx, c_idx)
            cell.value = row.get(name)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {1: 16, 2: 16, 3: 12, 4: 12, 5: 36, 6: 36, 7: 34, 8: 34, 9: 72, 10: 54, 11: 20}
    for col_idx, width in widths.items():
        wsr.column_dimensions[col_letter(col_idx)].width = width


def mark_data_sheet(ws, headers, reports):
    col_bin = headers["Storage Bin"]
    col_inv = headers["Inventariado"]
    col_date = headers["Data Fecho"]
    col_res = headers["Resumo SAP"]
    col_obs = headers["Observacoes Inventario"]
    by_loc_summary = defaultdict(list)
    by_loc_obs = defaultdict(list)
    by_loc_dt = {}
    for row in reports:
        loc = row["Storage Bin"]
        by_loc_summary[loc].append(f"{row['Product']}: {row['Resultado']}")
        by_loc_obs[loc].append(f"{row['Product']}: {row['Observacoes']}")
        by_loc_dt[loc] = row["Data Fecho"]
    for r in range(2, ws.max_row + 1):
        loc = to_text(ws.cell(r, col_bin).value)
        if loc in by_loc_summary:
            ws.cell(r, col_inv).value = "SIM"
            ws.cell(r, col_date).value = by_loc_dt[loc]
            ws.cell(r, col_res).value = " | ".join(by_loc_summary[loc])
            ws.cell(r, col_obs).value = " | ".join(by_loc_obs[loc])
            ws.cell(r, col_res).alignment = Alignment(wrap_text=True)
            ws.cell(r, col_obs).alignment = Alignment(wrap_text=True)
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = GREEN_FILL


def build_output_file():
    if not st.session_state.source_bytes:
        raise ValueError("Carrega primeiro o ficheiro de stocks.")
    wb = load_workbook(io.BytesIO(st.session_state.source_bytes))
    ws = wb["Data"]
    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    ensure_output_columns(ws, headers)
    headers = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }
    write_reports_sheet(wb, st.session_state.reports)
    mark_data_sheet(ws, headers, st.session_state.reports)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def articles_df():
    rows = []
    for product, rec in sorted(st.session_state.current_articles.items(), key=lambda x: x[0]):
        rows.append({
            "Artigo": product,
            "Descrição": rec["description"],
            "Qtd orig.": fmt_qty(rec["original_total"]),
            "Qtd cont.": fmt_qty(rec["counted_total"]),
            "HUs fechadas": "SIM" if rec["hus_closed"] else "NÃO",
            "Artigo fechado": "SIM" if rec["done"] else "NÃO",
        })
    return pd.DataFrame(rows)


def hu_df(product):
    rec = st.session_state.current_articles[product]
    all_hus = sorted(set(rec["original_hus"].keys()) | set(rec["counted_hus"].keys()))
    rows = []
    for hu in all_hus:
        orig_val = rec["original_hus"].get(hu, 0.0)
        counted_val = rec["counted_hus"].get(hu, None)
        diff_val = None if counted_val is None else float(counted_val) - float(orig_val)
        rows.append({
            "Código HU": hu,
            "Qtd original": orig_val,
            "Qtd contada": counted_val,
            "Diferença": diff_val,
        })
    return pd.DataFrame(rows)


def reports_df():
    rows = []
    for row in st.session_state.reports:
        rows.append({
            "Localização": row["Storage Bin"],
            "Artigo": row["Product"],
            "Qtd orig.": fmt_qty(row["Qtd Original"]),
            "Qtd cont.": fmt_qty(row["Qtd Contada"]),
            "HUs a criar": row["HUs a Criar"],
            "Resultado": row["Resultado"],
            "Ação SAP": row["Acao SAP"],
        })
    return pd.DataFrame(rows)


init_state()

st.markdown(
    """
    <style>
    .block-container {padding-top: 1.2rem; padding-bottom: 1rem;}
    .efacec-title {font-size: 2rem; font-weight: 800; color: #c41230; margin-bottom: 0.1rem;}
    .efacec-sub {color: #667085; margin-bottom: 1rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="efacec-title">EFACEC</div>', unsafe_allow_html=True)
st.markdown('<div class="efacec-sub">Inventário de Localizações — versão web interna em Streamlit</div>', unsafe_allow_html=True)

with st.container(border=True):
    c1, c2 = st.columns([2.5, 1.2])
    with c1:
        uploaded = st.file_uploader("Ficheiro de stocks", type=["xlsx", "xlsm", "xls"])
    with c2:
        st.text_input("Localização", key="location_input")

    b1, b2, b3 = st.columns([1, 1, 1.2])
    load_file_clicked = b1.button("Carregar ficheiro", use_container_width=True, type="primary")
    load_loc_clicked = b2.button("Carregar localização", use_container_width=True, type="primary")
    reset_clicked = b3.button("Nova localização", use_container_width=True)

if uploaded is not None:
    st.session_state.source_name = uploaded.name
    st.session_state.source_bytes = uploaded.getvalue()

if load_file_clicked:
    try:
        if not st.session_state.source_bytes:
            raise ValueError("Escolhe primeiro o ficheiro de stocks.")
        st.session_state.original_data = read_original_data_from_bytes(st.session_state.source_bytes)
        st.session_state.file_loaded = True
        st.session_state.closed_locations = set()
        st.session_state.reports = []
        reset_current_view(clear_location=False)
        st.success(
            f"Ficheiro carregado: {st.session_state.source_name} | Localizações: {len(st.session_state.original_data)}"
        )
    except Exception as e:
        st.error(str(e))

if load_loc_clicked:
    try:
        load_location(st.session_state.location_input)
        st.success(f"Localização carregada: {st.session_state.current_location}")
    except Exception as e:
        st.error(str(e))

if reset_clicked:
    reset_current_view(clear_location=True)
    st.info("Pronto para nova localização.")

left, right = st.columns([1.3, 1])

with left:
    with st.container(border=True):
        st.subheader("1. Artigos da localização")
        if st.session_state.current_articles:
            article_options = list(sorted(st.session_state.current_articles.keys()))
            default_idx = article_options.index(st.session_state.current_product) if st.session_state.current_product in article_options else 0
            selected_article = st.selectbox("Seleciona o artigo", article_options, index=default_idx)
            st.session_state.current_product = selected_article

            a1, a2 = st.columns([1, 1])
            if a1.button("Abrir HUs do artigo", use_container_width=True):
                try:
                    ensure_article_open(selected_article)
                    st.success(f"HUs abertas para o artigo {selected_article}.")
                except Exception as e:
                    st.error(str(e))

            if a2.button("Fechar artigo e passar ao próximo", use_container_width=True, type="primary"):
                try:
                    if st.session_state.current_product != selected_article:
                        st.session_state.current_product = selected_article
                    if st.session_state.hu_article_open != selected_article:
                        ensure_article_open(selected_article)
                    close_hus_for_current_article()
                    complete_article(selected_article)
                    nxt = next_open_article(selected_article)
                    st.session_state.current_product = nxt
                    st.success("Artigo fechado com sucesso.")
                except Exception as e:
                    st.error(str(e))

            st.dataframe(articles_df(), use_container_width=True, height=360, hide_index=True)
        else:
            st.info("Carrega o ficheiro e a localização para ver os artigos.")

with right:
    with st.container(border=True):
        st.subheader("2. Handling Units")
        product = st.session_state.current_product
        if product and product in st.session_state.current_articles:
            rec = st.session_state.current_articles[product]
            st.caption(f"Artigo atual: {product}")
            try:
                if st.session_state.hu_article_open == product:
                    df_hu = hu_df(product)
                    editor_key = f"hu_editor_{st.session_state.current_location}_{product}"
                    edited = st.data_editor(
                        df_hu,
                        use_container_width=True,
                        hide_index=True,
                        num_rows="dynamic",
                        key=editor_key,
                        disabled=["Qtd original", "Diferença"] if not rec["hus_closed"] else True,
                        column_config={
                            "Código HU": st.column_config.TextColumn("Código HU"),
                            "Qtd original": st.column_config.NumberColumn("Qtd original", format="%.3f"),
                            "Qtd contada": st.column_config.NumberColumn("Qtd contada", format="%.3f"),
                            "Diferença": st.column_config.NumberColumn("Diferença", format="%.3f"),
                        },
                    )

                    if not rec["hus_closed"]:
                        if st.button("Aplicar alterações das HUs", use_container_width=True):
                            new_counted = {}
                            for _, row in edited.iterrows():
                                hu = to_text(row["Código HU"])
                                if not hu:
                                    continue
                                counted = row["Qtd contada"]
                                new_counted[hu] = None if pd.isna(counted) else float(counted)
                            rec["counted_hus"] = new_counted
                            recalc_article(product)
                            rec["done"] = False
                            rec["hus_closed"] = False
                            st.success("Alterações das HUs aplicadas.")
                        c1, c2 = st.columns([1, 1])
                        with c1:
                            if st.button("Fechar HUs deste artigo", use_container_width=True):
                                try:
                                    new_counted = {}
                                    for _, row in edited.iterrows():
                                        hu = to_text(row["Código HU"])
                                        if not hu:
                                            continue
                                        counted = row["Qtd contada"]
                                        new_counted[hu] = None if pd.isna(counted) else float(counted)
                                    rec["counted_hus"] = new_counted
                                    close_hus_for_current_article()
                                    st.success("HUs fechadas com sucesso.")
                                except Exception as e:
                                    st.error(str(e))
                        with c2:
                            if st.button("Adicionar HU nova", use_container_width=True):
                                next_code = f"NOVA_HU_{len(rec['counted_hus']) + 1}"
                                rec["counted_hus"][next_code] = 0.0
                                rec["done"] = False
                                rec["hus_closed"] = False
                                st.success(f"HU adicionada: {next_code}.")
                    else:
                        st.info("As HUs deste artigo já estão fechadas.")
                else:
                    st.info("Clica em 'Abrir HUs do artigo' para começar a contar.")
            except Exception as e:
                st.error(str(e))
        else:
            st.info("Seleciona um artigo para ver as HUs.")

with st.container(border=True):
    st.subheader("3. Report da sessão")
    r1, r2 = st.columns([1, 1])
    if r1.button("Fechar localização atual", type="primary", use_container_width=True):
        try:
            close_location()
            st.success(f"Localização {st.session_state.current_location} fechada com sucesso.")
        except Exception as e:
            st.error(str(e))

    if st.session_state.reports:
        output_bytes = build_output_file()
        base_name = os.path.splitext(st.session_state.source_name or "inventario")[0]
        r2.download_button(
            "Guardar resultado Excel",
            data=output_bytes,
            file_name=f"{base_name}_inventario_resultado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        r2.button("Guardar resultado Excel", use_container_width=True, disabled=True)

    if st.session_state.reports:
        st.dataframe(reports_df(), use_container_width=True, height=280, hide_index=True)
        counts = defaultdict(int)
        for row in st.session_state.reports:
            counts[row["Storage Bin"]] += 1
        st.caption("Localizações fechadas nesta sessão: " + " | ".join(f"{loc}: {n} artigo(s)" for loc, n in sorted(counts.items())))
    else:
        st.info("Sem localizações fechadas nesta sessão.")
